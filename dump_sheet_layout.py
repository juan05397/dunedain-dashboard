import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\BRILLO\Downloads\Tarea\Dashboard ejemmplo2.xlsx", data_only=False)

def dump_sheet(ws, max_r=30, max_c=15):
    print(f"\n================ SHEET {ws.title} (up to {max_r}x{max_c}) ================")
    for r in range(1, min(max_r + 1, ws.max_row + 1)):
        row_vals = []
        for c in range(1, min(max_c + 1, ws.max_column + 1)):
            cell = ws.cell(r, c)
            val = cell.value
            if val is not None:
                if isinstance(val, str) and val.startswith("="):
                    row_vals.append(f"{cell.coordinate}:{val}")
                else:
                    row_vals.append(f"{cell.coordinate}:{repr(val)}")
        if row_vals:
            print(f"Row {r:02d}: " + " | ".join(row_vals))

dump_sheet(wb['datos'], max_r=30, max_c=15)
dump_sheet(wb['datos2'], max_r=30, max_c=15)
