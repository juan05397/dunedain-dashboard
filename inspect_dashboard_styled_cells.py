import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\BRILLO\Downloads\Tarea\Dashboard ejemmplo2.xlsx", data_only=False)
ws = wb['Dashboard']

print("================ MERGED CELLS ================")
print(list(ws.merged_cells.ranges))

print("\n================ POPULATED OR FORMATTED CELLS ================")
count = 0
for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(r, c)
        
        has_val = cell.value is not None
        has_fill = cell.fill and cell.fill.fill_type is not None
        has_border = cell.border and (cell.border.left.style or cell.border.right.style or cell.border.top.style or cell.border.bottom.style)
        has_font = cell.font and (cell.font.bold or cell.font.color and cell.font.color.value != '00000000')
        
        if has_val or has_fill or has_border or has_font:
            count += 1
            print(f"Cell {cell.coordinate}: Value={cell.value} | Fill={cell.fill.fill_type if cell.fill else None} | Color={cell.fill.start_color.rgb if cell.fill and cell.fill.fill_type else None} | FontBold={cell.font.bold if cell.font else None} | FontColor={cell.font.color.rgb if cell.font and cell.font.color else None}")
            if count > 50:
                print("Showing first 50 cells only...")
                break
    if count > 50:
        break
